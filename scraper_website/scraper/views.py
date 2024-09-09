from django.shortcuts import render

# Create your views here.
import pandas as pd
from django.shortcuts import render
import csv
import openpyxl
from bs4 import BeautifulSoup
import requests
from django.http import HttpResponse
from openpyxl import load_workbook

import csv
import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By

def index(request):
    if "GET" == request.method:
        return render(request, 'scraper/index.html', {})
    else:
        if(True):
            excel = openpyxl.Workbook()
            print(excel.sheetnames)
            sheet = excel.active
            sheet.title = "top rated shows"
            print(excel.sheetnames)
            sheet.append(["movie name", 'movie rank', "movie yar", "IMDB rating"])
            try:
                source = requests.get('https://www.imdb.com/chart/top/')
                source.raise_for_status()
                soup = BeautifulSoup(source.text, 'html.parser')

                movies = soup.find('tbody', class_="lister-list").find_all('tr')

                for movie in movies:
                    name = movie.find('td', class_="titleColumn").a.text

                    rank = movie.find('td', class_="titleColumn").get_text(strip=True).split('.')[0]

                    year = movie.find('td', class_="titleColumn").span.text.strip('()')

                    rating = movie.find('td', class_="ratingColumn imdbRating").strong.text

                    print(name, rank, year, rating)
                    sheet.append([name, rank, year, rating])



            except Exception as e:
                print(e.args)

            excel.save("top rated shows.xlsx")
            df = pd.read_excel('top rated shows.xlsx')
            html_table = df.to_html()
            return render(request, 'index.html', {'table': html_table})
        else:


            chromedriver_autoinstaller.install()
            driver = webdriver.Chrome()
            wait = WebDriverWait(driver, 20)
            driver.get("https://www.naukri.com/full-stack-developer-jobs?src=popular_roles_homepage_srch")

            count = 1000
            index, new_index, i = '0', 1, 0

            heading_xpath = '(//*[@class="jobTuple bgWhite br4 mb-8"])[' + index + ']/div/div/a'
            link_xpath = '(//*[@class="jobTuple bgWhite br4 mb-8"])[' + index + ']/div/div/a'
            subheading_xpath = '(//*[@class="jobTuple bgWhite br4 mb-8"])[' + index + ']/div/div/div/a'
            experience_xpath = '(//*[@class="jobTuple bgWhite br4 mb-8"])[' + index + ']/div/div/ul/li[1]/span'
            salary_xpath = '(//*[@class="jobTuple bgWhite br4 mb-8"])[' + index + ']/div/div/ul/li[2]/span'

            csv_file = open('Naukri_scrape.csv', 'a', encoding="utf-8", newline='')
            csv_writer = csv.writer(csv_file)
            csv_writer.writerow(['Heading', 'Sub Heading', 'Vacancy Link', 'Experience Needed', 'Salary'])

            while i < count:

                for j in range(20):
                    temp_index = str(new_index).zfill(2)
                    heading_xpath = heading_xpath.replace(index, temp_index)
                    link_xpath = link_xpath.replace(index, temp_index)
                    subheading_xpath = subheading_xpath.replace(index, temp_index)
                    experience_xpath = experience_xpath.replace(index, temp_index)
                    salary_xpath = salary_xpath.replace(index, temp_index)
                    index = str(new_index).zfill(2)
                    try:
                        heading = wait.until(EC.presence_of_element_located((By.XPATH, heading_xpath))).text
                        print(heading)
                    except:
                        heading = "NULL"
                    try:
                        link = wait.until(EC.presence_of_element_located((By.XPATH, link_xpath))).get_attribute('href')
                        print(link)
                    except:
                        link = "NULL"
                    try:
                        subheading = wait.until(EC.presence_of_element_located((By.XPATH, subheading_xpath))).text
                        print(subheading)
                    except:
                        subheading = "NULL"
                    try:
                        experience = wait.until(EC.presence_of_element_located((By.XPATH, experience_xpath))).text
                        print(experience)
                    except:
                        experience = "NULL"
                    try:
                        salary = wait.until(EC.presence_of_element_located((By.XPATH, salary_xpath))).text
                        print(salary)
                    except:
                        salary = "Not Disclosed"
                    new_index += 1
                    i += 1
                    print("--------------------------- " + str(i) + " ----------------------------------")

                    csv_writer.writerow([heading, subheading, link, experience, salary])
                    if i >= count:
                        break
                if i >= count:
                    break
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[text() = "Next"]'))).click()
                new_index = 1
            csv_file.close()
        excel_file = request.FILES["excel_file"]

        # you may put validations here to check extension or file size

        wb = openpyxl.load_workbook(excel_file)

        # getting all sheets
        sheets = wb.sheetnames
        print(sheets)

        # getting a particular sheet
        worksheet = wb["Sheet1"]
        print(worksheet)

        # getting active sheet
        active_sheet = wb.active
        print(active_sheet)

        # reading a cell
        print(worksheet["A1"].value)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
                print(cell.value)
            excel_data.append(row_data)

        return render(request, 'scrapper/index.html', {"excel_data":excel_data})

